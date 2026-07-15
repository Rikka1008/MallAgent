from pathlib import Path

from knowledge.ingestion.models import SourceDocument


def load_source_documents(source_dir: Path) -> list[SourceDocument]:
    """加载目录下所有支持的 RAG 源文件。

    目前支持 Markdown 和 Excel 问答表。你可以把这个函数理解成“知识库入口扫描器”：
    它会走遍目录，把能识别的文件转换成统一的 `SourceDocument`。
    """

    documents: list[SourceDocument] = []
    documents.extend(load_markdown_documents(source_dir))
    for path in sorted(source_dir.rglob("*.xlsx")):
        documents.extend(load_excel_qa_documents(path))
    return documents


def load_markdown_documents(source_dir: Path) -> list[SourceDocument]:
    """加载目录下所有 Markdown 文件。

    第一步只支持 `.md`，因为你的售后政策、FAQ、流程说明最适合先用 Markdown 管理。
    后续再扩展 PDF、Word、HTML，不要一开始把加载器写得过重。
    """

    documents: list[SourceDocument] = []
    for path in sorted(source_dir.rglob("*.md")):
        if path.name.upper() == "README.MD":
            continue
        documents.append(
            SourceDocument(
                text=path.read_text(encoding="utf-8"),
                metadata={
                    "source_path": str(path),
                    "source_name": path.name,
                    "source_type": "markdown",
                    "source_category": _guess_category(path),
                    "relative_path": str(path.relative_to(source_dir)),
                },
            )
        )
    return documents


def load_excel_qa_documents(file_path: Path) -> list[SourceDocument]:
    """加载 Excel 客服问答数据。

    你的 `网店客服回复数据集.xlsx` 是“问题/回复”两列，非常适合变成 FAQ 知识。
    每一行会变成一个 `SourceDocument`，后续再交给 splitter 切片。
    """

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 Excel 知识源需要安装 openpyxl。") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    question_index = _find_column(headers, ["问题", "question", "query"])
    answer_index = _find_column(headers, ["回复", "回答", "answer", "response"])
    if question_index is None or answer_index is None:
        raise ValueError("Excel FAQ 文件必须包含“问题”和“回复”两列。")

    documents: list[SourceDocument] = []
    for row_number, row in enumerate(rows[1:], start=2):
        question = _cell_to_text(row[question_index] if question_index < len(row) else None)
        answer = _cell_to_text(row[answer_index] if answer_index < len(row) else None)
        if not question or not answer:
            continue
        documents.append(
            SourceDocument(
                text=f"问题：{question}\n回复：{answer}",
                metadata={
                    "source_path": str(file_path),
                    "source_name": file_path.name,
                    "source_type": "excel",
                    "source_category": "faq",
                    "row_index": row_number,
                },
            )
        )
    return documents


def _find_column(headers: list[str], candidates: list[str]) -> int | None:
    """根据候选列名查找 Excel 列位置。"""

    lowered = [header.lower() for header in headers]
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered.index(candidate.lower())
    return None


def _cell_to_text(value) -> str:
    """把 Excel 单元格值转换成干净字符串。"""

    return "" if value is None else str(value).strip()


def _guess_category(path: Path) -> str:
    """根据目录名粗略判断知识分类。"""

    parts = {part.lower() for part in path.parts}
    for category in ["policies", "faq", "products", "logistics", "refunds", "mall_docs"]:
        if category in parts:
            return category
    return "raw_exports"
